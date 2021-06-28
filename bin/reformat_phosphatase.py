""" Extract phosphatase dataset from Huang et al. (2015)
"""

import os
import json
import time
import argparse
from collections import defaultdict
import re
import pandas as pd
import difflib

import rdkit
from rdkit import Chem
from rdkit.Chem.MolStandardize import rdMolStandardize
from rdkit.Chem.rdmolops import RemoveStereochemistry

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

import cirpy
import download_utils

PDB_CELL = "C4"
TBL_START = "C7"
TBL_END = "N20"

# As reported in paper
BINARY_THRESH = 0.2

# Use this to match the substrates across the sheets
MANUAL_MATCHES = {
    "PNPP": "p-nitrophenylphosphate (pNPP)",
    "NADP": "beta-Nicotinamide adenine dinucleotide phosphate (NADP)",
    "FMN": "riboflavin-5-phosphate (FMN)",
    "3'-5' ADP": "Adenosine 3',5'-diphosphate",
    "D-arabinose-5-phosphate": "arabinose 5-phosphate",
    "α-mannose-1-phosphate": "D-Mannose-2-P \xa0 \xa0",
}
START_CHIRAL = "^(L|D)-"

# Sheets not to iterate over
DIFF = {"Averages", "EFI-UniProt lookup ", " Layout"}


def extract_uniprot_to_efid(wb):
    """Extract the uniprot efi lookup"""
    mapping = {}
    for index, (efi, uniprot) in enumerate(wb["EFI-UniProt lookup "].iter_rows()):

        # Avoid header
        if index > 0:
            efi_text = str(efi.value).strip()
            uniprot_text = str(uniprot.value).strip()
            mapping[efi_text] = uniprot_text
    return mapping


def parse_sheet(wb, ws_name, mapping):
    """Parse the worksheet"""
    ws = wb[ws_name]
    uniprot_id = mapping[ws_name.strip()]
    pdb_id = ws[PDB_CELL].value

    dicts = []
    # Loop through rows
    for row in ws[TBL_START:TBL_END]:
        for cell in row:
            substrate = cell.comment.content.strip()
            value = cell.value
            if substrate.lower().strip() != "blank":
                my_dict = {
                    "SUBSTRATE": substrate,
                    "seq_id": uniprot_id,
                    "conversion": value,
                    "near_pdb": pdb_id,
                }
                dicts.append(my_dict)

    return dicts


def get_substrate_sheet_mapping(label_file, dicts):
    """Map the dict substrate names to the labels in the other file using
    fuzzy matching"""

    print("Mapping names between excel data sheets")
    substrate_label = pd.read_excel(
        label_file,
        engine="openpyxl",
    )[["Substrate", "Class"]]
    # Remove nan from this substrate label dataset (due to openpyxl)
    substrate_label = substrate_label[
        [isinstance(j, str) for j in substrate_label["Substrate"].values]
    ]
    match_dict = {}
    match_dict.update(MANUAL_MATCHES)
    unmatched = set([i["SUBSTRATE"] for i in dicts])
    match_options = set(substrate_label["Substrate"].values)

    unmatched = unmatched.difference(set(list(match_dict.keys())))
    match_options = match_options.difference(set(list(match_dict.values())))

    cutoff = 1.0
    # print(f"Start length of unmatched: {len(unmatched)}")
    # print(f"Start length of match_options: {len(match_options)}")

    while len(unmatched) > 0:
        for j in sorted(list(unmatched)):
            matches = difflib.get_close_matches(j, match_options, n=1, cutoff=cutoff)
            if len(matches) > 0:

                # Make sure we don't confuse chirality here
                if (
                    re.search(START_CHIRAL, matches[0])
                    and re.search(START_CHIRAL, j)
                    and matches[0][:2] != j[:2]
                ):
                    pass
                else:
                    match_dict[j] = matches[0]
                    match_options.remove(matches[0])
        # Update unmatched
        unmatched = unmatched.difference(set(list(match_dict.keys())))
        cutoff = max(cutoff - 0.01, 0)
        # print(f"New length of unmatched: {len(unmatched)}")
        # print(f"New length of match_options: {len(match_options)}")
        # print(f"New cutoff value: {cutoff}")

    print("Done mapping names between sheets")
    return match_dict


def main(rawdir, outdir, data_file, label_file, manually_annotated):
    """Main method"""
    wb = load_workbook(filename=data_file)

    mapping = extract_uniprot_to_efid(wb)

    sheets = wb.get_sheet_names()
    iter_items = set(sheets).difference(DIFF)

    dicts = []
    for sheet_name in iter_items:
        dicts.extend(parse_sheet(wb, sheet_name, mapping))

    raw_data = pd.DataFrame(dicts)
    label_dict_mapping = get_substrate_sheet_mapping(label_file, dicts)

    sub_list = set(pd.unique(raw_data["SUBSTRATE"]))

    new_mapping = defaultdict(
        lambda: {"pubchem": None, "cirpy": None, "manual": None, "final": None}
    )

    # Get cirpy
    for sub in sub_list:
        new_mapping[sub]["cirpy"] = cirpy.resolve(sub, representation="smiles")

    # Get pubchem
    temp_mapping = {k.replace("'", ""): k for k in sub_list}
    pubchem_mapping = download_utils.query_pubchem(
        set(temp_mapping.keys()),
        query_type="synonym",
        save_file="temp.txt",
        encoding="utf-8",
    )
    for sub in pubchem_mapping:
        sub_real = temp_mapping[sub]
        if sub_real in new_mapping:
            new_mapping[sub_real]["pubchem"] = pubchem_mapping[sub][0]

    # Get manual mapping
    df = pd.read_csv(manually_annotated, header=None, sep="\t")
    manual_mapping = dict(zip(df[0].values, df[1].values))
    for sub in manual_mapping:
        new_mapping[sub]["manual"] = manual_mapping[sub]

    priority = ["manual", "pubchem", "cirpy"]

    # K is new mapping
    for k in new_mapping:
        for obj in priority:
            if new_mapping[k][obj]:
                new_mapping[k]["final"] = new_mapping[k][obj]
                break

    # Standardize final molecules
    for k in new_mapping:
        smiles = new_mapping[k]["final"]
        m = Chem.MolFromSmiles(smiles)
        m = rdMolStandardize.Uncharger().uncharge(m)

        new_mapping[k]["final"] = Chem.MolToSmiles(m)

        # Make an achiral version
        RemoveStereochemistry(m)
        new_mapping[k]["final_achiral"] = Chem.MolToSmiles(m)

    # Look at how many of these smiles strings are redundant
    sub_mapping_ar = (
        pd.DataFrame(new_mapping)
        .transpose()
        .reset_index()
        .rename({"index": "subname"}, axis=1)
    )
    for k, v in sub_mapping_ar.groupby("final").size().items():
        if v != 1:
            redundant_smiles = k
            subs = sub_mapping_ar[sub_mapping_ar["final"] == redundant_smiles][
                "subname"
            ].values
            print(f"Smiles string: {k}")
            print(f"Corresponding compounds: {subs}")

    # Convert this to a dataframe
    df_export = (
        pd.DataFrame(new_mapping)
        .transpose()
        .reset_index()
        .rename({"index": "substrate"}, axis=1)
    )

    df_export = df_export[df_export["substrate"] != "substrate"].reset_index(drop=True)
    df_export["label_sheet_name"] = [
        label_dict_mapping[i] for i in df_export["substrate"]
    ]
    df_export.to_csv(os.path.join(rawdir, f"temp_annotations_{time.time()}.csv"))

    smiles = df_export["final"].values
    names = df_export["substrate"].values

    # Get uniprot seq ids
    raw_data = pd.DataFrame(dicts)
    uniprot_ids = pd.unique(raw_data["seq_id"])

    uniprot_file = os.path.join(rawdir, "phosphatase_sequence_file.txt")
    if os.path.exists(uniprot_file):
        os.remove(uniprot_file)
    download_utils.download_uniprot_uniparc_list(uniprot_file, uniprot_ids)

    df = pd.read_csv(uniprot_file, sep="\t")

    id_to_seq = dict(zip(df["Entry"].values.tolist(), df["Sequence"].values.tolist()))

    # Make chiral and achiral versions of this dataset
    for chiral_name, chiral_key in zip(
        ["chiral", "achiral"], ["final", "final_achiral"]
    ):
        out_dicts_regr = []
        out_dicts_bin = []
        for sub, seq, conv in raw_data[["SUBSTRATE", "seq_id", "conversion"]].values:
            conv = float(conv)
            smiles_final = new_mapping[sub][chiral_key]

            aa_seq = id_to_seq[seq]
            binary_conv = 1 if conv > BINARY_THRESH else 0

            out_dict_regr = {
                "SEQ": aa_seq,
                "SUBSTRATES": smiles_final,
                "Conversion": conv,
            }
            out_dict_bin = {
                "SEQ": aa_seq,
                "SUBSTRATES": smiles_final,
                "Conversion": binary_conv,
            }

            out_dicts_regr.append(out_dict_regr)
            out_dicts_bin.append(out_dict_bin)

        bin_df = pd.DataFrame(out_dicts_bin)
        reg_df = pd.DataFrame(out_dicts_regr)

        # Take the maximum value for equivalent seq,smiles pairs
        bin_df = bin_df.groupby(["SEQ", "SUBSTRATES"]).max().reset_index()
        reg_df = reg_df.groupby(["SEQ", "SUBSTRATES"]).max().reset_index()

        out_regr = os.path.join(outdir, f"phosphatase_{chiral_name}.csv")
        out_bin = os.path.join(outdir, f"phosphatase_{chiral_name}_binary.csv")
        bin_df.to_csv(out_bin)
        reg_df.to_csv(out_regr)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--rawdir", type=str, help="Raw data directory", default="data/raw"
    )
    parser.add_argument("--outdir", type=str, help="outdir", default="data/processed")
    parser.add_argument(
        "--data-file",
        type=str,
        help="Data file as input (sd02).xlsx",
        default="data/raw/phosphatase/pnas.1423570112.sd02.xlsx",
    )
    parser.add_argument(
        "--label-file",
        type=str,
        help="Data file as input (sd01).xlsx",
        default="data/raw/phosphatase/pnas.1423570112.sd01.xlsx",
    )
    parser.add_argument(
        "--manually-annotated",
        type=str,
        help="Manually annotated common name to smiles string, tab separated",
        default="data/raw/phosphatase/smiles.dat",
    )
    args = parser.parse_args()

    rawdir = args.rawdir
    outdir = args.outdir
    data_file = args.data_file
    label_file = args.label_file
    manually_annotated = args.manually_annotated

    os.makedirs(outdir, exist_ok=True)

    main(rawdir, outdir, data_file, label_file, manually_annotated)
